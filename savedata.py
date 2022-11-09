import openpyxl as xl
import WLalg
import time
import networkx as nx
import WLalgShevarshidze as WLV2
def generategraphs(graphgenerator, range):
    graphs = dict()
    for i in range:
        graph = graphgenerator(i)
        graphs.update({i:graph})
    return graphs
def savedata(filename, graphs:dict, graphisomorthismfunction):
    data = dict()
    for size in graphs.keys():
        G = graphs[size]
        H = G.copy()
        data.update({size: []})
        for i in range(20):
            start = time.time()
            graphisomorthismfunction(G, H)
            end = time.time()
            data[size].append(end - start)
        print('finished ', size)

    wb = xl.Workbook()
    ws = wb.active
    ws.title = 'data'
    i = 1
    for key in data.keys():
        ws.cell(row=i, column=1, value=key)
        for j in range(20):
            ws.cell(row=i, column=j + 2, value=data[key][j])
        i += 1
    wb.save(filename=filename)
graphs = generategraphs(nx.path_graph, range(30, 1510, 10))
savedata('data1.xlsx', graphs, WLalg.wlalg)
graphs = generategraphs(nx.complete_graph, range(30, 610, 10))
savedata('data4.xlsx', graphs, WLalg.wlalg)

